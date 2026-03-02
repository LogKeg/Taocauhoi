"""
Grading API endpoints for OMR and handwritten answer sheets.

Note: The helper functions are currently in app.main and will be
refactored into app.grading module in a future update.
"""
import io
import json
from typing import List

from fastapi import APIRouter, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from app.core import ANSWER_TEMPLATES

router = APIRouter(prefix="/api", tags=["grading"])


# Forward references to grading functions in main.py
# These will be replaced with direct imports after full refactoring


def _grade_handwritten_sheet(
    image_bytes: bytes,
    answer_key: List[str],
    num_questions: int = 30,
    valid_answers: List[str] = None
) -> dict:
    """
    Grade a handwritten answer sheet using OCR.

    Note: This is a placeholder implementation. Full OCR-based grading
    requires external services (Google Vision API, Tesseract, etc.)
    """
    import cv2
    import numpy as np

    if valid_answers is None:
        valid_answers = ['A', 'B', 'C', 'D', 'E']

    try:
        # Decode image
        np_arr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)

        if img is None:
            return {"error": "Không thể đọc ảnh"}

        # Placeholder: Return empty result - real implementation needs OCR
        return {
            "detected_answers": [],
            "correct": 0,
            "wrong": 0,
            "blank": num_questions,
            "total": num_questions,
            "details": [],
            "note": "OCR handwritten grading not fully implemented. Please use OMR grading for multiple choice."
        }

    except Exception as e:
        return {"error": f"Lỗi xử lý ảnh: {str(e)}"}


def _get_grading_functions():
    """Lazy import of grading functions from main module."""
    from app import main
    return {
        'grade_single_sheet': main._grade_single_sheet,
        'grade_mixed_format_sheet': main._grade_mixed_format_sheet,
        'preprocess_omr_image': main._preprocess_omr_image,
        'detect_bubbles_grid_based': main._detect_bubbles_grid_based,
        'group_bubbles_to_questions_improved': main._group_bubbles_to_questions_improved,
        'extract_student_info_ocr': main._extract_student_info_ocr,
        'detect_template_from_image': main._detect_template_from_image,
        'parse_answer_key_for_template': main._parse_answer_key_for_template,
        'grade_handwritten_sheet': _grade_handwritten_sheet,
    }


@router.post("/grade-sheets")
async def grade_answer_sheets(
    files: List[UploadFile],
    template_type: str = Form("IKSC_BENJAMIN"),
    answer_key: str = Form(None),
    answer_file: UploadFile = None,
    auto_detect_template: bool = Form(True)
):
    """Grade multiple OMR answer sheets.

    If auto_detect_template=True (default), the system will automatically detect
    the exam type (IKSC/IKLC) and level (Benjamin, Cadet, etc.) from the answer sheet
    and retrieve the corresponding answer key from the answer file.
    """
    funcs = _get_grading_functions()

    # Read answer file content if provided
    answer_file_content = None
    answer_file_ext = None
    if answer_file and answer_file.filename:
        answer_file_content = await answer_file.read()
        answer_file_ext = answer_file.filename.lower().split(".")[-1]

    # Cache parsed answers for each template
    answers_cache = {}

    def get_answers_for_template(tpl_type: str) -> List[str]:
        """Get answers for a template, using cache."""
        if tpl_type in answers_cache:
            return answers_cache[tpl_type]

        template = ANSWER_TEMPLATES.get(tpl_type)
        if not template:
            return []

        num_questions = template["questions"]
        answers = []

        if answer_file_content and answer_file_ext:
            try:
                answers = funcs['parse_answer_key_for_template'](
                    answer_file_content, answer_file_ext, tpl_type
                )
            except Exception:
                pass

        if not answers and answer_key:
            try:
                answers = json.loads(answer_key)
                if isinstance(answers, str):
                    answers = list(answers.upper())
                answers = [str(a).upper() for a in answers]
            except json.JSONDecodeError:
                if "," in answer_key:
                    answers = [a.strip().upper() for a in answer_key.split(",")]
                else:
                    answers = list(answer_key.upper().replace(" ", ""))

        answers_cache[tpl_type] = answers
        return answers

    # Handle AUTO mode
    if template_type == "AUTO":
        auto_detect_template = True
        template_type = "IKSC_BENJAMIN"

    # Validate default template
    default_template = ANSWER_TEMPLATES.get(template_type)
    if not default_template:
        raise HTTPException(status_code=400, detail=f"Invalid template type: {template_type}")

    # If not auto-detecting, validate answer key first
    if not auto_detect_template:
        answers = get_answers_for_template(template_type)
        num_questions = default_template["questions"]
        if len(answers) < num_questions:
            raise HTTPException(
                status_code=400,
                detail=f"Missing answers. Need {num_questions}, got {len(answers)}"
            )

    results = []
    all_scores = []

    for file in files:
        if not file.filename:
            continue

        ext = file.filename.lower().split(".")[-1]
        supported_image_formats = ["jpg", "jpeg", "png", "bmp", "tiff"]

        if ext == "pdf":
            try:
                import fitz
                content = await file.read()
                pdf_doc = fitz.open(stream=content, filetype="pdf")

                for page_num in range(len(pdf_doc)):
                    page = pdf_doc[page_num]
                    mat = fitz.Matrix(2.0, 2.0)
                    pix = page.get_pixmap(matrix=mat)
                    img_bytes = pix.tobytes("png")

                    page_filename = f"{file.filename}_page_{page_num + 1}"

                    actual_template = template_type
                    if auto_detect_template:
                        student_info = funcs['extract_student_info_ocr'](img_bytes)
                        detected = student_info.get("detected_template", "")
                        if detected and detected in ANSWER_TEMPLATES:
                            actual_template = detected

                    answers = get_answers_for_template(actual_template)
                    tpl = ANSWER_TEMPLATES.get(actual_template, default_template)

                    if len(answers) < tpl["questions"]:
                        results.append({
                            "filename": page_filename,
                            "detected_template": actual_template,
                            "error": f"No answers found for {tpl['name']}. Please check the answer file."
                        })
                        continue

                    if tpl.get("mixed_format"):
                        result = funcs['grade_mixed_format_sheet'](img_bytes, answers, actual_template)
                    else:
                        result = funcs['grade_single_sheet'](img_bytes, answers, actual_template)

                    if "error" in result:
                        results.append({
                            "filename": page_filename,
                            "detected_template": actual_template,
                            "error": result["error"]
                        })
                    else:
                        result["filename"] = page_filename
                        result["detected_template"] = actual_template
                        results.append(result)
                        all_scores.append(result["score"])

                pdf_doc.close()
            except Exception as e:
                results.append({
                    "filename": file.filename,
                    "error": f"PDF read error: {str(e)}"
                })
            continue

        elif ext not in supported_image_formats:
            results.append({
                "filename": file.filename,
                "error": "Unsupported file format"
            })
            continue

        try:
            content = await file.read()

            actual_template = template_type
            if auto_detect_template:
                detected_info = funcs['detect_template_from_image'](content)
                detected = detected_info.get("detected_template", "")

                if detected and detected in ANSWER_TEMPLATES:
                    actual_template = detected
                else:
                    try:
                        result_temp = funcs['preprocess_omr_image'](content)
                        if result_temp[0] is not None:
                            _, gray, binary = result_temp

                            rows, _ = funcs['detect_bubbles_grid_based'](gray, binary, "IKSC_BENJAMIN")
                            questions_iksc = funcs['group_bubbles_to_questions_improved'](rows, "IKSC_BENJAMIN")
                            num_iksc = len(questions_iksc)

                            rows_iklc, _ = funcs['detect_bubbles_grid_based'](gray, binary, "IKLC_STUDENT")
                            questions_iklc = funcs['group_bubbles_to_questions_improved'](rows_iklc, "IKLC_STUDENT")
                            num_iklc = len(questions_iklc)

                            num_questions_detected = max(num_iksc, num_iklc)
                            is_iklc_layout = num_iklc > num_iksc

                            if num_questions_detected >= 45 or is_iklc_layout:
                                level = detected_info.get('detected_level', 'BENJAMIN')
                                if level not in ['BENJAMIN', 'CADET', 'JUNIOR', 'STUDENT']:
                                    level = 'STUDENT'
                                actual_template = f"IKLC_{level}"
                                if actual_template not in ANSWER_TEMPLATES:
                                    actual_template = "IKLC_STUDENT"
                            elif num_questions_detected >= 25:
                                actual_template = f"IKSC_{detected_info.get('detected_level', 'BENJAMIN')}"
                                if actual_template not in ANSWER_TEMPLATES:
                                    actual_template = "IKSC_BENJAMIN"
                            elif num_questions_detected >= 20:
                                actual_template = "IKSC_PRE_ECOLIER"

                            if detected_info.get("detected_contest"):
                                contest = detected_info["detected_contest"]
                                level = detected_info.get("detected_level", "")
                                if level:
                                    test_template = f"{contest}_{level}"
                                    if test_template in ANSWER_TEMPLATES:
                                        actual_template = test_template
                    except:
                        pass

            answers = get_answers_for_template(actual_template)
            tpl = ANSWER_TEMPLATES.get(actual_template, default_template)

            if len(answers) < tpl["questions"]:
                results.append({
                    "filename": file.filename,
                    "detected_template": actual_template,
                    "error": f"No answers found for {tpl['name']}. Please check the answer file."
                })
                continue

            if tpl.get("mixed_format"):
                result = funcs['grade_mixed_format_sheet'](content, answers, actual_template)
            else:
                result = funcs['grade_single_sheet'](content, answers, actual_template)

            if "error" in result:
                results.append({
                    "filename": file.filename,
                    "detected_template": actual_template,
                    "error": result["error"]
                })
            else:
                result["filename"] = file.filename
                result["detected_template"] = actual_template
                results.append(result)
                all_scores.append(result["score"])
        except Exception as e:
            results.append({
                "filename": file.filename,
                "error": f"Processing error: {str(e)}"
            })

    summary = {
        "total_sheets": len(results),
        "graded": len(all_scores),
        "errors": len(results) - len(all_scores),
        "average_score": round(sum(all_scores) / len(all_scores), 2) if all_scores else 0,
        "highest": max(all_scores) if all_scores else 0,
        "lowest": min(all_scores) if all_scores else 0
    }

    return {
        "ok": True,
        "auto_detect": auto_detect_template,
        "results": results,
        "summary": summary
    }


@router.post("/grade-sheets/export")
async def export_grading_results(
    results: str = Form(...),
    template_type: str = Form("IKSC_BENJAMIN")
):
    """Export grading results to Excel."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    try:
        data = json.loads(results)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid data format")

    num_questions = 30
    for result in data:
        if "details" in result and result["details"]:
            num_questions = len(result["details"])
            break
        elif "total" in result:
            num_questions = result["total"]
            break

    if num_questions == 30:
        template = ANSWER_TEMPLATES.get(template_type, ANSWER_TEMPLATES["IKSC_BENJAMIN"])
        num_questions = template["questions"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grading Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["#", "Name", "Class", "DOB", "ID", "School", "Score", "Correct", "Wrong", "Blank"]
    for i in range(1, num_questions + 1):
        headers.append(f"Q{i}")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, result in enumerate(data, 2):
        if "error" in result:
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=result.get("filename", ""))
            ws.cell(row=row_idx, column=7, value="Error: " + result["error"])
            continue

        student_info = result.get("student_info", {})

        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align
        ws.cell(row=row_idx, column=2, value=student_info.get("full_name", "")).alignment = left_align
        ws.cell(row=row_idx, column=3, value=student_info.get("class", "")).alignment = center_align
        ws.cell(row=row_idx, column=4, value=student_info.get("dob", "")).alignment = center_align
        ws.cell(row=row_idx, column=5, value=student_info.get("id_no", "")).alignment = center_align
        ws.cell(row=row_idx, column=6, value=student_info.get("school_name", "")).alignment = left_align
        ws.cell(row=row_idx, column=7, value=result.get("score", 0)).alignment = center_align
        ws.cell(row=row_idx, column=8, value=result.get("correct", 0)).alignment = center_align
        ws.cell(row=row_idx, column=9, value=result.get("wrong", 0)).alignment = center_align
        ws.cell(row=row_idx, column=10, value=result.get("blank", 0)).alignment = center_align

        details = result.get("details", [])
        for detail in details:
            col = 10 + detail["q"]
            cell = ws.cell(row=row_idx, column=col, value=detail.get("student", ""))
            cell.alignment = center_align
            cell.border = thin_border

            if detail["status"] == "correct":
                cell.fill = correct_fill
            elif detail["status"] in ["wrong", "invalid"]:
                cell.fill = wrong_fill

    answer_row = len(data) + 2
    ws.cell(row=answer_row, column=2, value="ANSWER KEY").font = Font(bold=True)

    if data and "details" in data[0]:
        for detail in data[0]["details"]:
            col = 10 + detail["q"]
            cell = ws.cell(row=answer_row, column=col, value=detail.get("correct", ""))
            cell.alignment = center_align
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 6
    ws.column_dimensions["I"].width = 6
    ws.column_dimensions["J"].width = 10

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=grading_results_{template_type}.xlsx"}
    )


@router.post("/grade-handwritten")
async def grade_handwritten_sheets(
    files: List[UploadFile],
    answer_key: str = Form(...),
    num_questions: int = Form(30),
    valid_answers: str = Form("A,B,C,D,E"),
    scoring_correct: float = Form(1.0),
    scoring_wrong: float = Form(0.0),
    scoring_blank: float = Form(0.0)
):
    """Grade multiple handwritten answer sheets using OCR."""
    funcs = _get_grading_functions()

    try:
        if answer_key.startswith('['):
            keys = json.loads(answer_key)
        else:
            keys = [k.strip().upper() for k in answer_key.split(',')]
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid answer key format")

    valid_list = [v.strip().upper() for v in valid_answers.split(',')]

    results = []

    for file in files:
        try:
            image_bytes = await file.read()

            result = funcs['grade_handwritten_sheet'](
                image_bytes,
                keys,
                num_questions,
                valid_list
            )

            if "error" not in result:
                score = (
                    result["correct"] * scoring_correct +
                    result["wrong"] * scoring_wrong +
                    result["blank"] * scoring_blank
                )
                result["score"] = round(score, 2)

            result["filename"] = file.filename
            results.append(result)

        except Exception as e:
            results.append({
                "filename": file.filename,
                "error": str(e)
            })

    valid_results = [r for r in results if "error" not in r]
    summary = {
        "total_sheets": len(files),
        "successful": len(valid_results),
        "failed": len(files) - len(valid_results)
    }

    if valid_results:
        scores = [r["score"] for r in valid_results]
        summary["average_score"] = round(sum(scores) / len(scores), 2)
        summary["highest"] = max(scores)
        summary["lowest"] = min(scores)

    return {
        "ok": True,
        "results": results,
        "summary": summary
    }


@router.post("/grade-handwritten/export")
async def export_handwritten_results(
    results: str = Form(...),
    num_questions: int = Form(30)
):
    """Export handwritten grading results to Excel."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    try:
        data = json.loads(results)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid data format")

    for result in data:
        if "details" in result and result["details"]:
            num_questions = len(result["details"])
            break

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Handwritten Grading Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    low_conf_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["#", "Filename", "Score", "Correct", "Wrong", "Blank"]
    for i in range(1, num_questions + 1):
        headers.append(f"Q{i}")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, result in enumerate(data, 2):
        if "error" in result:
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=result.get("filename", ""))
            ws.cell(row=row_idx, column=3, value="Error: " + result["error"])
            continue

        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align
        ws.cell(row=row_idx, column=2, value=result.get("filename", "")).alignment = center_align
        ws.cell(row=row_idx, column=3, value=result.get("score", 0)).alignment = center_align
        ws.cell(row=row_idx, column=4, value=result.get("correct", 0)).alignment = center_align
        ws.cell(row=row_idx, column=5, value=result.get("wrong", 0)).alignment = center_align
        ws.cell(row=row_idx, column=6, value=result.get("blank", 0)).alignment = center_align

        details = result.get("details", [])
        for detail in details:
            col = 6 + detail["q"]
            cell = ws.cell(row=row_idx, column=col, value=detail.get("student", ""))
            cell.alignment = center_align
            cell.border = thin_border

            if detail["status"] == "correct":
                cell.fill = correct_fill
            elif detail["status"] in ["wrong", "invalid"]:
                cell.fill = wrong_fill

            confidence = detail.get("confidence", 1.0)
            if confidence < 0.5 and detail["status"] != "blank":
                cell.fill = low_conf_fill

    answer_row = len(data) + 2
    ws.cell(row=answer_row, column=2, value="ANSWER KEY").font = Font(bold=True)

    if data and "details" in data[0]:
        for detail in data[0]["details"]:
            col = 6 + detail["q"]
            cell = ws.cell(row=answer_row, column=col, value=detail.get("correct", ""))
            cell.alignment = center_align
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 10

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=handwritten_grading_results.xlsx"}
    )
