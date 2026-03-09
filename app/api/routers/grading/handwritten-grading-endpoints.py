"""
Handwritten answer sheet grading endpoints.

POST /api/grade-handwritten - Grade multiple handwritten answer sheets using OCR.
POST /api/grade-handwritten/export - Export handwritten grading results to Excel.
"""
import io
import json
from typing import List

from fastapi import APIRouter, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from app.api.routers.grading.helpers import _get_grading_functions

router = APIRouter(prefix="/api", tags=["grading"])


@router.post("/grade-handwritten")
async def grade_handwritten_sheets(
    files: List[UploadFile],
    answer_key: str = Form(...),
    num_questions: int = Form(30),
    valid_answers: str = Form("A,B,C,D,E"),
    scoring_correct: float = Form(1.0),
    scoring_wrong: float = Form(0.0),
    scoring_blank: float = Form(0.0)
):
    """Grade multiple handwritten answer sheets using OCR."""
    funcs = _get_grading_functions()

    try:
        if answer_key.startswith('['):
            keys = json.loads(answer_key)
        else:
            keys = [k.strip().upper() for k in answer_key.split(',')]
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid answer key format")

    valid_list = [v.strip().upper() for v in valid_answers.split(',')]

    results = []

    for file in files:
        try:
            image_bytes = await file.read()

            result = funcs['grade_handwritten_sheet'](
                image_bytes,
                keys,
                num_questions,
                valid_list
            )

            if "error" not in result:
                score = (
                    result["correct"] * scoring_correct +
                    result["wrong"] * scoring_wrong +
                    result["blank"] * scoring_blank
                )
                result["score"] = round(score, 2)

            result["filename"] = file.filename
            results.append(result)

        except Exception as e:
            results.append({
                "filename": file.filename,
                "error": str(e)
            })

    valid_results = [r for r in results if "error" not in r]
    summary = {
        "total_sheets": len(files),
        "successful": len(valid_results),
        "failed": len(files) - len(valid_results)
    }

    if valid_results:
        scores = [r["score"] for r in valid_results]
        summary["average_score"] = round(sum(scores) / len(scores), 2)
        summary["highest"] = max(scores)
        summary["lowest"] = min(scores)

    return {
        "ok": True,
        "results": results,
        "summary": summary
    }


@router.post("/grade-handwritten/export")
async def export_handwritten_results(
    results: str = Form(...),
    num_questions: int = Form(30)
):
    """Export handwritten grading results to Excel."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    try:
        data = json.loads(results)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid data format")

    for result in data:
        if "details" in result and result["details"]:
            num_questions = len(result["details"])
            break

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Handwritten Grading Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    low_conf_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["#", "Filename", "Score", "Correct", "Wrong", "Blank"]
    for i in range(1, num_questions + 1):
        headers.append(f"Q{i}")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, result in enumerate(data, 2):
        if "error" in result:
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=result.get("filename", ""))
            ws.cell(row=row_idx, column=3, value="Error: " + result["error"])
            continue

        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align
        ws.cell(row=row_idx, column=2, value=result.get("filename", "")).alignment = center_align
        ws.cell(row=row_idx, column=3, value=result.get("score", 0)).alignment = center_align
        ws.cell(row=row_idx, column=4, value=result.get("correct", 0)).alignment = center_align
        ws.cell(row=row_idx, column=5, value=result.get("wrong", 0)).alignment = center_align
        ws.cell(row=row_idx, column=6, value=result.get("blank", 0)).alignment = center_align

        details = result.get("details", [])
        for detail in details:
            col = 6 + detail["q"]
            cell = ws.cell(row=row_idx, column=col, value=detail.get("student", ""))
            cell.alignment = center_align
            cell.border = thin_border

            if detail["status"] == "correct":
                cell.fill = correct_fill
            elif detail["status"] in ["wrong", "invalid"]:
                cell.fill = wrong_fill

            confidence = detail.get("confidence", 1.0)
            if confidence < 0.5 and detail["status"] != "blank":
                cell.fill = low_conf_fill

    answer_row = len(data) + 2
    ws.cell(row=answer_row, column=2, value="ANSWER KEY").font = Font(bold=True)

    if data and "details" in data[0]:
        for detail in data[0]["details"]:
            col = 6 + detail["q"]
            cell = ws.cell(row=answer_row, column=col, value=detail.get("correct", ""))
            cell.alignment = center_align
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 10

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=handwritten_grading_results.xlsx"}
    )
